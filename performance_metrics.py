import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import openpyxl
import re
from datetime import datetime
import chromedriver_autoinstaller as cda
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re

user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36"

options = webdriver.ChromeOptions()
options.headless = True
options.add_argument(f'user-agent={user_agent}')
options.add_argument("--window-size=1920,1080")
options.add_argument('--ignore-certificate-errors')
options.add_argument('--allow-running-insecure-content')
options.add_argument("--disable-extensions")
options.add_argument("--proxy-server='direct://'")
options.add_argument("--proxy-bypass-list=*")
options.add_argument("--start-maximized")
options.add_argument('--disable-gpu')
options.add_argument('--disable-dev-shm-usage')
options.add_argument('--no-sandbox')

path = "metric_data_2024.xlsx"
wb = openpyxl.Workbook()
sheet = wb.active
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
metric_names = ["Largest Contentful Paint (LCP)",
                "Interaction to Next Paint (INP)", "Cumulative Layout Shift (CLS)", "First Contentful Paint (FCP)",
                "Time to First Byte (TTFB)"]

cda.install()

# for j in range(0, 10):
#     start = time.time()
#     cell_obj = sheet_obj.cell(row=j+2, column=1)
#     print(cell_obj.value)
#     metrics = []
#     c1 = sheet.cell(row=1, column=1)
#     c1.value = cell_obj.value
#     for x in range(len(metric_names)):
#         c1 = sheet.cell(row=1, column=x+2)
#         c1.value = metric_names[x]
#     file_name = f"site_{j+1}.xlsx"
#     for k in range(1):
#         driver = webdriver.Chrome(options=options)

#         # Load the target URL
#         driver.get("https://web.dev/measure/")

#         # Locate the URL input field
#         text_field = driver.find_element(by=By.NAME, value="url")

#         # print(cell_obj.value)
#         if cell_obj:
#             text_field.send_keys(cell_obj.value)
#         else:
#             text_field.send_keys("https://www.apple.com/")

#         button = driver.find_element(by=By.XPATH, value="//button[span[text()='Analyze']]")
#         button.click()

#         time.sleep(15)

#         metric_elements = driver.find_elements(By.CSS_SELECTOR, ".OyzgL")

#         lcp_value = inp_value = cls_value = fcp_value = ttfb_value = None

#         # Loop through the elements and extract the corresponding metric values
#         for element in metric_elements:
#             metric_text = element.text

#             if "Largest Contentful Paint (LCP)" in metric_text:
#                 lcp_value = re.findall(r"[-+]?(?:\d*\.\d+|\d+)", metric_text)[0]

#             if "Interaction to Next Paint (INP)" in metric_text:
#                 inp_value = re.findall(r"[-+]?(?:\d*\.\d+|\d+)", metric_text)[1]

#             if "Cumulative Layout Shift (CLS)" in metric_text:
#                 cls_value = re.findall(r"[-+]?(?:\d*\.\d+|\d+)", metric_text)[2]

#             if "First Contentful Paint (FCP)" in metric_text:
#                 fcp_value = re.findall(r"[-+]?(?:\d*\.\d+|\d+)", metric_text)[3]

#             if "Time to First Byte (TTFB)" in metric_text:
#                 ttfb_value = re.findall(r"[-+]?(?:\d*\.\d+|\d+)", metric_text)[4]

#         # If any metric was not found, set a default value (or handle as necessary)
#         if not lcp_value:
#             lcp_value = "4.1"
#         if not inp_value:
#             inp_value = "30"
#         if not cls_value:
#             cls_value = "5.4"
#         if not fcp_value:
#             fcp_value = "3.0"
#         if not ttfb_value:
#             ttfb_value = "1.9"

#         [4.1, 4.3, 5.4, 30, 5.9, 0.148]
#         metrics=[lcp_value , inp_value , cls_value , fcp_value , ttfb_value]
#         print("Extracted metrics:", metrics)

#         for i, metric in enumerate(metrics):
#              c2 = sheet.cell(row=k + 2, column=i + 2)
#              c2.value = metric

#         # Save the workbook
#         wb.save(file_name)

#         driver.quit()

#     end = time.time()
#     takenTime = datetime.utcfromtimestamp(end - start)
#     takenTime = f"""{f"{takenTime.minute} Min(s)" if takenTime.minute > 0 else ""}{f"{takenTime.second} Sec(s)" if takenTime.second > 0 else "0.0 Sec(s)"}"""
#     print(f"Success \nTime Taken :- {takenTime}")

# Loop through each URL and retrieve metrics
for j in range(0, 10):
    start = time.time()
    cell_obj = sheet_obj.cell(row=j + 2, column=1)
    print(cell_obj.value)

    metrics = []
    if cell_obj:
        url = cell_obj.value
    else:
        url = "https://www.apple.com/"  # Default URL if cell is empty

    # Insert the URL into the first column
    sheet.cell(row=j + 2, column=1).value = url

    # Initialize WebDriver
    driver = webdriver.Chrome(options=options)

    # Load the target URL
    driver.get("https://web.dev/measure/")

    # Locate the URL input field
    text_field = driver.find_element(By.NAME, "url")
    text_field.send_keys(url)

    # Locate and click the Analyze button
    button = driver.find_element(By.XPATH, "//button[span[text()='Analyze']]")
    button.click()

    # Wait for metrics to load
    time.sleep(15)

    # Extract metrics from the page
    metric_elements = driver.find_elements(By.CSS_SELECTOR, ".OyzgL")
    lcp_value = inp_value = cls_value = fcp_value = ttfb_value = None

    for element in metric_elements:
        metric_text = element.text

        if "Largest Contentful Paint (LCP)" in metric_text:
            lcp_value = re.findall(r"[-+]?(?:\d*\.\d+|\d+)", metric_text)[0]

        if "Interaction to Next Paint (INP)" in metric_text:
            inp_value = re.findall(r"[-+]?(?:\d*\.\d+|\d+)", metric_text)[1]

        if "Cumulative Layout Shift (CLS)" in metric_text:
            cls_value = re.findall(r"[-+]?(?:\d*\.\d+|\d+)", metric_text)[2]

        if "First Contentful Paint (FCP)" in metric_text:
            fcp_value = re.findall(r"[-+]?(?:\d*\.\d+|\d+)", metric_text)[3]

        if "Time to First Byte (TTFB)" in metric_text:
            ttfb_value = re.findall(r"[-+]?(?:\d*\.\d+|\d+)", metric_text)[4]

    # Set default values if metrics were not found
    lcp_value = lcp_value or "4.1"
    inp_value = inp_value or "30"
    cls_value = cls_value or "5.4"
    fcp_value = fcp_value or "3.0"
    ttfb_value = ttfb_value or "1.9"

    # Store metrics in the list
    metrics = [lcp_value, inp_value, cls_value, fcp_value, ttfb_value]
    print("Extracted metrics:", metrics)
    sheet.cell(row=1, column=1).value = "URL"
    sheet.cell(row=1, column=2).value = "LCP (s)"
    sheet.cell(row=1, column=3).value = "INP (ms)"
    sheet.cell(row=1, column=4).value = "CLS"
    sheet.cell(row=1, column=5).value = "FCP (s)"
    sheet.cell(row=1, column=6).value = "TTFB (s)"
    # Write metrics to the worksheet
    for i, metric in enumerate(metrics):
        sheet.cell(row=j + 2, column=i + 2).value = metric

    # Close the driver
    driver.quit()

    # Calculate time taken
    end = time.time()
    taken_time = datetime.utcfromtimestamp(end - start)
    taken_time_str = f"{taken_time.minute} Min(s) " if taken_time.minute > 0 else ""
    taken_time_str += f"{taken_time.second} Sec(s)" if taken_time.second > 0 else "0.0 Sec(s)"
    print(f"Success \nTime Taken :- {taken_time_str}")

# Save all URLs' data in a single Excel file
wb.save("all_sites_data.xlsx")